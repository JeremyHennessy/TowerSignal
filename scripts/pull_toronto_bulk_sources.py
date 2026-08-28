from __future__ import annotations

import argparse
import csv
import io
import json
import shutil
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.parse import urlencode
from urllib.request import Request, urlopen

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - installed in workflow
    load_workbook = None

ROOT = Path(__file__).resolve().parents[1]

TORONTO_CKAN = "https://ckan0.cf.opendata.inter.prod-toronto.ca/api/3/action"
USER_AGENT = "TowerSignal-Toronto-Warehouse/0.1 (+https://github.com/JeremyHennessy/TowerSignal)"
TORONTO_OPEN_DATA_LICENSE = "Open Government Licence - Toronto"
ONTARIO_OPEN_DATA_LICENSE = "Open Government Licence - Ontario"

KEYWORDS = [
    "cooling tower",
    "cooling towers",
    "chiller",
    "chillers",
    "condenser water",
    "evaporative condenser",
    "water treatment",
    "cooling water",
    "legionella",
    "chemical feed",
    "mechanical",
    "hvac",
    "boiler",
]

OPEN_DATA_SPECS = [
    {
        "key": "tobids_awarded_contracts",
        "slug": "tobids-awarded-contracts",
        "title": "Toronto Bids Awarded Contracts",
        "resource_hint": "Awarded Contracts",
        "mode": "all",
    },
    {
        "key": "capital_project_pipeline",
        "slug": "capital-project-pipeline",
        "title": "Capital Projects Pipeline",
        "resource_hint": "Capital Project Pipeline",
        "mode": "all",
    },
    {
        "key": "development_pipeline",
        "slug": "development-pipeline",
        "title": "Development Pipeline",
        "resource_hint": "Development Pipeline",
        "mode": "all",
    },
    {
        "key": "chemtrac_2024",
        "slug": "chemical-tracking-chemtrac",
        "title": "Chemical Tracking (ChemTrac)",
        "resource_hint": "Chemtrac Data 2024",
        "mode": "all",
    },
    {
        "key": "apartment_building_evaluation",
        "slug": "apartment-building-evaluation",
        "title": "Apartment Building Evaluation",
        "resource_hint": "Apartment Building Evaluation",
        "mode": "all",
        "quality_note": "City Open Data has an active notice that this dataset is incomplete/not updating correctly; retain source warning.",
    },
    {
        "key": "affordable_housing_pipeline",
        "slug": "upcoming-and-recently-completed-affordable-housing-units",
        "title": "Affordable Housing Pipeline",
        "resource_hint": "Affordable Rental Housing Pipeline",
        "mode": "all",
    },
]

METADATA_ONLY_SPECS = [
    {
        "key": "address_points",
        "slug": "address-points-municipal-toronto-one-address-repository",
        "title": "Address Points (Municipal) - Toronto One Address Repository",
        "reason": "Large geospatial spine; use targeted joins rather than duplicating full city geometry in git.",
    },
    {
        "key": "building_outlines",
        "slug": "topographic-mapping-building-outlines",
        "title": "Topographic Mapping - Building Outlines",
        "reason": "Large geospatial source; retain metadata and add targeted spatial extraction later.",
    },
]

BPS_2024_URL = (
    "https://data.ontario.ca/dataset/5e976319-9769-4f77-aab9-d170e0131efe/"
    "resource/b6b569cf-5d3a-4b96-9972-ea428c316278/download/2024_final_data_set.xlsx"
)

ACCESS_ENVIRONMENT_BASES = [
    "https://ws.lioservices.lrc.gov.on.ca/arcgis1071a/rest/services/Access_Environment/Access_Environment_Map/MapServer",
    "https://ws.lioservices.lrc.gov.on.ca/arcgis2/rest/services/Access_Environment/Access_Environment_Map/MapServer",
]
ACCESS_ENVIRONMENT_LAYERS = {
    "environmental_compliance_approvals": 1,
    "environmental_activity_sector_registrations": 3,
    "permits_to_take_water": 5,
}
TORONTO_BBOX = (-79.6393, 43.5810, -79.1150, 43.8554)

SOURCE_INVENTORY_REVIEW = [
    {
        "key": "application_information_centre_documents",
        "source_url": "https://www.toronto.ca/city-government/planning-development/application-information-centre/",
        "status": "DISCOVERED_DOCUMENT_HEAVY_NOT_BULK_INGESTED_YET",
        "reason": "Requires application/document enumeration and PDF text extraction; source is daily-current.",
    },
    {
        "key": "construction_act_certificates",
        "source_url": "https://www.ontario.ca/laws/statute/90c30",
        "status": "DISCOVERED_REUSE_REVIEW_REQUIRED",
        "reason": "Certificates are public records, but bulk publication platforms are third-party; verify reuse/automation terms before republishing scraped records.",
    },
    {
        "key": "tssa_bpv",
        "source_url": "https://www.tssa.org/guideline-access-public-information",
        "status": "DISCOVERED_PAID_OR_REQUEST_BASED_BULK",
        "reason": "TSSA offers address inquiries and bulk database products; not an unauthenticated open bulk feed.",
    },
    {
        "key": "toronto_aerial_2025_8cm",
        "source_url": "https://gis.toronto.ca/arcgis/rest/services/basemap/cot_ortho_2025_color_8cm/MapServer",
        "status": "DISCOVERED_SERVICE_METADATA_ONLY",
        "reason": "Imagery tiles are a computer-vision enrichment source, not suitable for duplicating wholesale into git.",
    },
]


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")


def request_bytes(url: str, *, timeout: int = 90, retries: int = 4) -> bytes:
    last_error: Exception | None = None
    for attempt in range(retries):
        try:
            request = Request(
                url,
                headers={
                    "User-Agent": USER_AGENT,
                    "Accept": "application/json,text/csv,application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*;q=0.8",
                    "Accept-Language": "en-CA,en;q=0.9",
                },
            )
            with urlopen(request, timeout=timeout) as response:
                return response.read()
        except (HTTPError, URLError, TimeoutError) as exc:
            last_error = exc
            if attempt + 1 < retries:
                time.sleep(1.5 * (attempt + 1))
    raise RuntimeError(f"Failed to retrieve {url}: {last_error}")


def request_json(url: str) -> dict[str, Any]:
    payload = json.loads(request_bytes(url).decode("utf-8"))
    if not isinstance(payload, dict):
        raise RuntimeError(f"Expected JSON object from {url}")
    return payload


def ckan_action(action: str, params: dict[str, Any]) -> dict[str, Any]:
    url = f"{TORONTO_CKAN}/{action}?{urlencode(params)}"
    payload = request_json(url)
    if payload.get("success") is not True:
        raise RuntimeError(f"Toronto CKAN action {action} failed: {payload!r}")
    result = payload.get("result")
    if not isinstance(result, dict):
        raise RuntimeError(f"Toronto CKAN action {action} returned no result object")
    return result


def resolve_package(slug: str, title: str) -> dict[str, Any]:
    try:
        return ckan_action("package_show", {"id": slug})
    except Exception:
        search = ckan_action("package_search", {"q": f'title:"{title}"', "rows": 20})
        results = search.get("results") or []
        exact = [item for item in results if str(item.get("title") or "").strip().lower() == title.lower()]
        if not exact:
            raise RuntimeError(f"Could not resolve Toronto Open Data package: {title!r} / {slug!r}")
        return exact[0]


def choose_resource(package: dict[str, Any], hint: str | None) -> dict[str, Any]:
    resources = [item for item in (package.get("resources") or []) if isinstance(item, dict)]
    if not resources:
        raise RuntimeError(f"Package {package.get('title')} has no resources")
    hint_norm = (hint or "").strip().lower()
    if hint_norm:
        exact = [r for r in resources if str(r.get("name") or "").strip().lower() == hint_norm]
        if exact:
            return exact[0]
        contains = [r for r in resources if hint_norm in str(r.get("name") or "").lower()]
        if contains:
            contains.sort(key=lambda r: (0 if r.get("datastore_active") else 1, str(r.get("name") or "")))
            return contains[0]
    datastore = [r for r in resources if r.get("datastore_active")]
    if datastore:
        return datastore[0]
    preferred = [
        r for r in resources
        if str(r.get("format") or "").upper() in {"CSV", "JSON", "XLSX", "XLS"}
    ]
    if preferred:
        return preferred[0]
    return resources[0]


def fetch_datastore_rows(resource_id: str, *, max_rows: int = 100_000) -> tuple[list[dict[str, Any]], int]:
    rows: list[dict[str, Any]] = []
    offset = 0
    limit = 5000
    total: int | None = None
    while total is None or offset < total:
        result = ckan_action(
            "datastore_search",
            {"resource_id": resource_id, "limit": limit, "offset": offset},
        )
        if total is None:
            total = int(result.get("total") or 0)
            if total > max_rows:
                raise RuntimeError(
                    f"Resource {resource_id} has {total:,} rows, exceeding repository snapshot cap {max_rows:,}"
                )
        batch = result.get("records") or []
        if not isinstance(batch, list):
            raise RuntimeError("CKAN datastore records were not a list")
        rows.extend(item for item in batch if isinstance(item, dict))
        if not batch:
            break
        offset += len(batch)
    return rows, int(total or 0)


def csv_rows(data: bytes) -> list[dict[str, Any]]:
    text = data.decode("utf-8-sig", errors="replace")
    return [dict(row) for row in csv.DictReader(io.StringIO(text))]


def xlsx_rows(data: bytes) -> list[dict[str, Any]]:
    if load_workbook is None:
        raise RuntimeError("openpyxl is required for XLSX sources")
    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    rows_out: list[dict[str, Any]] = []
    for sheet in workbook.worksheets:
        iterator = sheet.iter_rows(values_only=True)
        header = None
        for row in iterator:
            values = [str(v).strip() if v is not None else "" for v in row]
            if sum(bool(v) for v in values) >= 3:
                header = values
                break
        if not header:
            continue
        header = [value or f"column_{idx+1}" for idx, value in enumerate(header)]
        for row in iterator:
            values = list(row)
            if not any(value is not None and str(value).strip() for value in values):
                continue
            record = {}
            for idx, name in enumerate(header):
                value = values[idx] if idx < len(values) else None
                if isinstance(value, datetime):
                    value = value.isoformat()
                record[name] = value
            rows_out.append(record)
    return rows_out


def fetch_resource_rows(resource: dict[str, Any], *, max_rows: int = 100_000) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    if resource.get("datastore_active"):
        rows, total = fetch_datastore_rows(str(resource["id"]), max_rows=max_rows)
        return rows, {"retrieval": "CKAN_DATASTORE", "reported_total": total}
    url = str(resource.get("url") or "")
    if not url:
        raise RuntimeError(f"Resource {resource.get('name')} has no URL")
    data = request_bytes(url)
    if len(data) > 30_000_000:
        raise RuntimeError(f"Resource {resource.get('name')} exceeds 30 MB snapshot cap")
    fmt = str(resource.get("format") or "").upper()
    if fmt == "CSV" or url.lower().endswith(".csv"):
        rows = csv_rows(data)
    elif fmt == "XLSX" or url.lower().endswith(".xlsx"):
        rows = xlsx_rows(data)
    elif fmt == "JSON" or url.lower().endswith(".json"):
        payload = json.loads(data.decode("utf-8"))
        if isinstance(payload, list):
            rows = [item for item in payload if isinstance(item, dict)]
        elif isinstance(payload, dict):
            candidate = payload.get("records") or payload.get("data") or payload.get("result")
            if not isinstance(candidate, list):
                raise RuntimeError(f"Could not locate JSON rows in {url}")
            rows = [item for item in candidate if isinstance(item, dict)]
        else:
            raise RuntimeError(f"Unexpected JSON shape from {url}")
    else:
        raise RuntimeError(f"Unsupported tabular resource format {fmt!r}: {url}")
    if len(rows) > max_rows:
        raise RuntimeError(f"Resource {resource.get('name')} parsed to {len(rows):,} rows over cap {max_rows:,}")
    return rows, {"retrieval": f"RESOURCE_{fmt or 'UNKNOWN'}", "download_bytes": len(data), "reported_total": len(rows)}


def stringify_record(record: dict[str, Any]) -> str:
    return " | ".join(str(value) for value in record.values() if value not in (None, ""))


def keyword_matches(record: dict[str, Any]) -> list[str]:
    text = stringify_record(record).lower()
    return sorted({keyword for keyword in KEYWORDS if keyword in text})


def write_json(path: Path, payload: Any, *, pretty: bool = False) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if pretty:
        path.write_text(json.dumps(payload, indent=2, ensure_ascii=False, default=str), encoding="utf-8")
    else:
        path.write_text(json.dumps(payload, separators=(",", ":"), ensure_ascii=False, default=str), encoding="utf-8")


def snapshot_open_dataset(spec: dict[str, Any], output_dir: Path, retrieved_at: str) -> dict[str, Any]:
    package = resolve_package(spec["slug"], spec["title"])
    resource = choose_resource(package, spec.get("resource_hint"))
    rows, retrieval = fetch_resource_rows(resource)
    relevant = []
    for row in rows:
        matches = keyword_matches(row)
        if matches:
            enriched = dict(row)
            enriched["_towersignal_keyword_matches"] = matches
            relevant.append(enriched)
    payload = {
        "metadata": {
            "key": spec["key"],
            "title": package.get("title"),
            "package_name": package.get("name"),
            "package_id": package.get("id"),
            "portal_url": f"https://open.toronto.ca/dataset/{package.get('name')}/",
            "license": package.get("license_title") or TORONTO_OPEN_DATA_LICENSE,
            "retrieved_at": retrieved_at,
            "resource": {
                "id": resource.get("id"),
                "name": resource.get("name"),
                "format": resource.get("format"),
                "last_modified": resource.get("last_modified"),
                "url": resource.get("url"),
                "datastore_active": bool(resource.get("datastore_active")),
            },
            "retrieval": retrieval,
            "quality_note": spec.get("quality_note"),
            "row_count": len(rows),
            "keyword_match_row_count": len(relevant),
            "keywords": KEYWORDS,
        },
        "rows": rows,
        "keyword_matches": relevant,
    }
    write_json(output_dir / f"{spec['key']}.json", payload)
    return payload["metadata"]


def snapshot_metadata_only(spec: dict[str, Any], output_dir: Path, retrieved_at: str) -> dict[str, Any]:
    package = resolve_package(spec["slug"], spec["title"])
    metadata = {
        "key": spec["key"],
        "title": package.get("title"),
        "package_name": package.get("name"),
        "package_id": package.get("id"),
        "portal_url": f"https://open.toronto.ca/dataset/{package.get('name')}/",
        "license": package.get("license_title") or TORONTO_OPEN_DATA_LICENSE,
        "retrieved_at": retrieved_at,
        "mode": "METADATA_ONLY",
        "reason": spec["reason"],
        "resources": [
            {
                "id": resource.get("id"),
                "name": resource.get("name"),
                "format": resource.get("format"),
                "last_modified": resource.get("last_modified"),
                "url": resource.get("url"),
                "datastore_active": bool(resource.get("datastore_active")),
            }
            for resource in (package.get("resources") or [])
            if isinstance(resource, dict)
        ],
    }
    write_json(output_dir / f"{spec['key']}_metadata.json", metadata, pretty=True)
    return metadata


def bps_toronto_candidate(record: dict[str, Any]) -> bool:
    text = stringify_record(record).lower()
    explicit_places = ("toronto", "etobicoke", "scarborough", "north york", "east york")
    return any(place in text for place in explicit_places)


def snapshot_bps(output_dir: Path, retrieved_at: str) -> dict[str, Any]:
    data = request_bytes(BPS_2024_URL)
    rows = xlsx_rows(data)
    candidates = [row for row in rows if bps_toronto_candidate(row)]
    relevant = []
    for row in candidates:
        matches = keyword_matches(row)
        if matches:
            enriched = dict(row)
            enriched["_towersignal_keyword_matches"] = matches
            relevant.append(enriched)
    payload = {
        "metadata": {
            "key": "ontario_bps_energy_2024",
            "title": "Ontario Broader Public Sector Energy Use and GHG Emissions - 2024",
            "source_url": BPS_2024_URL,
            "catalogue_url": "https://data.ontario.ca/dataset/energy-use-and-greenhouse-gas-emissions-for-the-broader-public-sector",
            "license": ONTARIO_OPEN_DATA_LICENSE,
            "retrieved_at": retrieved_at,
            "download_bytes": len(data),
            "source_row_count": len(rows),
            "toronto_candidate_row_count": len(candidates),
            "keyword_match_row_count": len(relevant),
            "candidate_contract": "Row text contains Toronto, Etobicoke, Scarborough, North York, or East York; candidate universe only, not a cooling-tower assertion.",
        },
        "toronto_candidates": candidates,
        "keyword_matches": relevant,
    }
    write_json(output_dir / "ontario_bps_energy_2024.json", payload)
    return payload["metadata"]


def arcgis_service_base() -> str:
    errors = []
    for base in ACCESS_ENVIRONMENT_BASES:
        try:
            payload = request_json(f"{base}?f=pjson")
            if payload.get("layers"):
                return base
        except Exception as exc:
            errors.append(f"{base}: {exc}")
    raise RuntimeError("Could not reach Access Environment ArcGIS service: " + " | ".join(errors))


def arcgis_layer_rows(base: str, layer_id: int) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    layer_meta = request_json(f"{base}/{layer_id}?f=pjson")
    page_size = min(int(layer_meta.get("maxRecordCount") or 2000), 2000)
    offset = 0
    features: list[dict[str, Any]] = []
    while True:
        params = {
            "where": "1=1",
            "geometry": ",".join(str(value) for value in TORONTO_BBOX),
            "geometryType": "esriGeometryEnvelope",
            "inSR": 4326,
            "spatialRel": "esriSpatialRelIntersects",
            "outFields": "*",
            "returnGeometry": "true",
            "outSR": 4326,
            "resultOffset": offset,
            "resultRecordCount": page_size,
            "f": "json",
        }
        payload = request_json(f"{base}/{layer_id}/query?{urlencode(params)}")
        if payload.get("error"):
            raise RuntimeError(f"ArcGIS layer {layer_id} query error: {payload['error']}")
        batch = payload.get("features") or []
        if not isinstance(batch, list):
            raise RuntimeError(f"ArcGIS layer {layer_id} features not a list")
        features.extend(item for item in batch if isinstance(item, dict))
        if len(batch) < page_size:
            break
        offset += len(batch)
        if len(features) > 50_000:
            raise RuntimeError(f"Access Environment layer {layer_id} exceeded 50,000 Toronto-bbox features")
    return features, {
        "layer_id": layer_id,
        "name": layer_meta.get("name"),
        "max_record_count": layer_meta.get("maxRecordCount"),
        "field_names": [field.get("name") for field in (layer_meta.get("fields") or []) if isinstance(field, dict)],
    }


def snapshot_access_environment(rights_review_dir: Path, retrieved_at: str) -> dict[str, Any]:
    base = arcgis_service_base()
    layer_summaries = {}
    for key, layer_id in ACCESS_ENVIRONMENT_LAYERS.items():
        features, meta = arcgis_layer_rows(base, layer_id)
        relevant = []
        for feature in features:
            attrs = feature.get("attributes") or {}
            if isinstance(attrs, dict):
                matches = keyword_matches(attrs)
                if matches:
                    relevant.append({"feature": feature, "_towersignal_keyword_matches": matches})
        payload = {
            "metadata": {
                "key": key,
                "source_service": base,
                "retrieved_at": retrieved_at,
                "scope": "TORONTO_APPROXIMATE_BOUNDING_BOX",
                "bbox_wgs84": TORONTO_BBOX,
                "rights_status": "PUBLIC_ACCESS_REUSE_REVIEW_REQUIRED_BEFORE_GIT_REPUBLICATION",
                "feature_count": len(features),
                "keyword_match_feature_count": len(relevant),
                **meta,
            },
            "features": features,
            "keyword_matches": relevant,
        }
        write_json(rights_review_dir / f"{key}.json", payload)
        layer_summaries[key] = payload["metadata"]
    return {
        "key": "ontario_access_environment",
        "source_service": base,
        "retrieved_at": retrieved_at,
        "rights_status": "PUBLIC_ACCESS_REUSE_REVIEW_REQUIRED_BEFORE_GIT_REPUBLICATION",
        "layers": layer_summaries,
    }


def build(output_dir: Path) -> dict[str, Any]:
    retrieved_at = utc_now()
    temp_dir = output_dir.parent / f".{output_dir.name}.tmp"
    if temp_dir.exists():
        shutil.rmtree(temp_dir)
    temp_dir.mkdir(parents=True, exist_ok=True)
    open_dir = temp_dir / "open_licensed"
    rights_review_dir = temp_dir / "rights_review"
    open_dir.mkdir(parents=True, exist_ok=True)
    rights_review_dir.mkdir(parents=True, exist_ok=True)

    source_summaries = []
    failures = []

    for spec in OPEN_DATA_SPECS:
        try:
            source_summaries.append(snapshot_open_dataset(spec, open_dir, retrieved_at))
        except Exception as exc:
            failures.append({"key": spec["key"], "error": str(exc)})

    for spec in METADATA_ONLY_SPECS:
        try:
            source_summaries.append(snapshot_metadata_only(spec, open_dir, retrieved_at))
        except Exception as exc:
            failures.append({"key": spec["key"], "error": str(exc)})

    try:
        source_summaries.append(snapshot_bps(open_dir, retrieved_at))
    except Exception as exc:
        failures.append({"key": "ontario_bps_energy_2024", "error": str(exc)})

    try:
        access_environment = snapshot_access_environment(rights_review_dir, retrieved_at)
    except Exception as exc:
        access_environment = {
            "key": "ontario_access_environment",
            "status": "FETCH_FAILED",
            "error": str(exc),
            "retrieved_at": retrieved_at,
        }
        failures.append({"key": "ontario_access_environment", "error": str(exc)})

    inventory = {
        "schema_version": "toronto-warehouse-0.1",
        "generated_at": retrieved_at,
        "jurisdiction": "TORONTO_ON",
        "purpose": "Raw/near-raw source warehouse for TowerSignal Toronto discovery and commercial intelligence. No Toronto priority score.",
        "open_licensed_sources": source_summaries,
        "rights_review_sources": [access_environment],
        "discovered_not_bulk_ingested": SOURCE_INVENTORY_REVIEW,
        "failures": failures,
        "contracts": {
            "source_absence": "No record in any source is not evidence that a property lacks a cooling tower.",
            "keyword_hits": "Keyword matches are discovery aids only and do not establish equipment identity or commercial need.",
            "rights_review": "Data under rights_review is fetched for source evaluation/artifact inspection and must not be republished to git until reuse rights are confirmed.",
        },
    }
    write_json(temp_dir / "source_inventory.json", inventory, pretty=True)

    core_keys = {
        "tobids_awarded_contracts",
        "capital_project_pipeline",
        "development_pipeline",
        "chemtrac_2024",
        "ontario_bps_energy_2024",
    }
    succeeded = {item.get("key") for item in source_summaries}
    missing_core = sorted(core_keys - succeeded)
    if missing_core:
        raise RuntimeError(f"Toronto warehouse missing required core sources: {missing_core}; failures={failures}")

    if output_dir.exists():
        shutil.rmtree(output_dir)
    temp_dir.rename(output_dir)
    print(json.dumps({
        "open_source_count": len(source_summaries),
        "failure_count": len(failures),
        "rights_review_layer_count": len((access_environment.get("layers") or {})) if isinstance(access_environment, dict) else 0,
    }, indent=2))
    return inventory


def main() -> None:
    parser = argparse.ArgumentParser(description="Pull broad Toronto/Ontario source warehouse for TowerSignal")
    parser.add_argument("--output", type=Path, default=ROOT / "data/toronto/warehouse/current")
    args = parser.parse_args()
    build(args.output)


if __name__ == "__main__":
    main()
