from __future__ import annotations

import argparse
import io
import json
import re
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.parse import urlencode
from urllib.request import Request, urlopen

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
CKAN = "https://data.ontario.ca/api/3/action"
USER_AGENT = "TowerSignal-Toronto-Warehouse/0.1 (+https://github.com/JeremyHennessy/TowerSignal)"
ONTARIO_LICENSE = "Open Government Licence - Ontario"
TORONTO_CITIES = {"TORONTO", "ETOBICOKE", "NORTH YORK", "SCARBOROUGH", "EAST YORK", "YORK"}

SOURCES = [
    {
        "key": "ontario_ewrb_large_buildings",
        "package": "energy-and-water-usage-of-large-buildings-in-ontario",
        "title": "Energy and water usage of large buildings in Ontario",
        "mode": "all_english_xlsx",
    },
    {
        "key": "ontario_environmental_compliance_reports",
        "package": "environmental-compliance-reports",
        "title": "Environmental Compliance Reports",
        "mode": "all_tabular",
    },
]


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")


def request_bytes(url: str, *, timeout: int = 120, retries: int = 4) -> bytes:
    last_error: Exception | None = None
    for attempt in range(retries):
        try:
            request = Request(url, headers={"User-Agent": USER_AGENT, "Accept": "application/json,*/*;q=0.8"})
            with urlopen(request, timeout=timeout) as response:
                return response.read()
        except (HTTPError, URLError, TimeoutError) as exc:
            last_error = exc
            if attempt + 1 < retries:
                time.sleep(1.5 * (attempt + 1))
    raise RuntimeError(f"Failed to retrieve {url}: {last_error}")


def request_json(url: str) -> dict[str, Any]:
    payload = json.loads(request_bytes(url).decode("utf-8"))
    if not isinstance(payload, dict):
        raise RuntimeError(f"Expected JSON object from {url}")
    return payload


def ckan(action: str, params: dict[str, Any]) -> dict[str, Any]:
    payload = request_json(f"{CKAN}/{action}?{urlencode(params)}")
    if payload.get("success") is not True or not isinstance(payload.get("result"), dict):
        raise RuntimeError(f"Ontario CKAN {action} failed: {payload!r}")
    return payload["result"]


def package_show(package_id: str) -> dict[str, Any]:
    return ckan("package_show", {"id": package_id})


def normalize_licence_label(value: Any) -> str:
    text = str(value or "")
    text = re.sub(r"[\u2010\u2011\u2012\u2013\u2014\u2015\u2212]", "-", text)
    text = re.sub(r"\s+", " ", text).strip().lower()
    return text


def clean_header(value: Any, index: int) -> str:
    text = re.sub(r"\s+", " ", str(value or "")).strip()
    return text or f"column_{index + 1}"


def sheet_rows(sheet) -> list[dict[str, Any]]:
    iterator = sheet.iter_rows(values_only=True)
    header: list[str] | None = None
    for row in iterator:
        values = list(row)
        nonempty = sum(1 for value in values if value is not None and str(value).strip())
        if nonempty >= 3:
            header = [clean_header(value, idx) for idx, value in enumerate(values)]
            break
    if not header:
        return []
    output: list[dict[str, Any]] = []
    for row in iterator:
        values = list(row)
        if not any(value is not None and str(value).strip() for value in values):
            continue
        record: dict[str, Any] = {}
        for idx, name in enumerate(header):
            value = values[idx] if idx < len(values) else None
            if isinstance(value, datetime):
                value = value.isoformat()
            record[name] = value
        output.append(record)
    return output


def xlsx_rows(data: bytes) -> list[dict[str, Any]]:
    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    rows: list[dict[str, Any]] = []
    for sheet in workbook.worksheets:
        rows.extend(sheet_rows(sheet))
    return rows


def csv_rows(data: bytes) -> list[dict[str, Any]]:
    import csv
    text = io.StringIO(data.decode("utf-8-sig", errors="replace"))
    return [dict(row) for row in csv.DictReader(text)]


def city_fields(record: dict[str, Any]) -> list[str]:
    return [
        key for key in record
        if re.sub(r"[^a-z0-9]", "", key.lower()) in {
            "city", "municipality", "municipalityname", "municipalname", "community", "towncity"
        }
    ]


def address_fields(record: dict[str, Any]) -> list[str]:
    output = []
    for key in record:
        normalized = re.sub(r"[^a-z0-9]", "", key.lower())
        if "address" in normalized and "email" not in normalized:
            output.append(key)
    return output


def Toronto_value(value: Any) -> bool:
    return str(value or "").strip().upper() in TORONTO_CITIES


def row_is_toronto(record: dict[str, Any]) -> tuple[bool, str]:
    cities = city_fields(record)
    if cities:
        return any(Toronto_value(record.get(field)) for field in cities), "EXACT_CITY_OR_MUNICIPALITY_FIELD"
    # Do not use generic text search as Toronto membership. If no city field exists,
    # a Toronto postal-code prefix is acceptable only when an address/location-like field carries it.
    for field in address_fields(record):
        text = str(record.get(field) or "").upper()
        if re.search(r"\bM\d[A-Z]\s*\d[A-Z]\d\b", text):
            return True, "TORONTO_POSTAL_CODE_IN_ADDRESS_FIELD"
    return False, "NO_DETERMINISTIC_TORONTO_FIELD"


def resource_year(resource: dict[str, Any]) -> int | None:
    text = " ".join(str(resource.get(key) or "") for key in ("name", "description", "url"))
    years = re.findall(r"\b(20\d{2})\b", text)
    return max((int(year) for year in years), default=None)


def resource_summary(resource: dict[str, Any]) -> dict[str, Any]:
    return {
        "id": resource.get("id"),
        "name": resource.get("name"),
        "format": resource.get("format"),
        "url": resource.get("url"),
        "last_modified": resource.get("last_modified"),
        "language": resource.get("language"),
        "year": resource_year(resource),
    }


def environmental_resource_key(resource: dict[str, Any]) -> tuple[int | None, str]:
    return resource_year(resource), re.sub(r"\s+", " ", str(resource.get("name") or "").strip().lower())


def select_preferred_environmental_resources(resources: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Choose one English tabular representation for each year/category."""
    english = [
        item for item in resources
        if "french" not in str(item.get("name") or "").lower()
        and not str(item.get("language") or "").lower().startswith("fr")
        and "_fr." not in str(item.get("url") or "").lower()
    ]
    grouped: dict[tuple[int | None, str], list[dict[str, Any]]] = {}
    for item in english:
        grouped.setdefault(environmental_resource_key(item), []).append(item)
    selected: list[dict[str, Any]] = []
    excluded = [item for item in resources if item not in english]
    for key in sorted(grouped, key=lambda item: (item[0] or 0, item[1])):
        variants = grouped[key]
        variants.sort(key=lambda item: (
            str(item.get("format") or "").upper() == "XLSX",
            str(item.get("last_modified") or ""),
            str(item.get("id") or ""),
        ))
        selected.append(variants[-1])
        excluded.extend(variants[:-1])
    return selected, excluded


def select_resources(package: dict[str, Any], mode: str) -> list[dict[str, Any]]:
    resources = [item for item in (package.get("resources") or []) if isinstance(item, dict)]
    if mode == "all_english_xlsx":
        selected = []
        for item in resources:
            fmt = str(item.get("format") or "").upper()
            name = str(item.get("name") or "").lower()
            language = str(item.get("language") or "").lower()
            if fmt != "XLSX":
                continue
            if "french" in name or language.startswith("fr"):
                continue
            selected.append(item)
        return selected
    if mode == "all_tabular":
        tabular = [
            item for item in resources
            if str(item.get("format") or "").upper() in {"XLSX", "CSV"}
        ]
        return select_preferred_environmental_resources(tabular)[0]
    raise RuntimeError(f"Unknown selection mode: {mode}")


def normalize_existing_environmental_report(path: Path) -> dict[str, int]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    metadata = payload.get("metadata") or {}
    resources = [item for item in (metadata.get("resources") or []) if isinstance(item, dict)]
    selected, excluded = select_preferred_environmental_resources(resources)
    selected_ids = {str(item.get("id") or "") for item in selected}
    rows = [
        row for row in (payload.get("toronto_rows") or [])
        if isinstance(row, dict) and str(row.get("_towersignal_source_resource_id") or "") in selected_ids
    ]
    metadata["resources"] = selected
    metadata["resource_count"] = len(selected)
    metadata["toronto_row_count"] = len(rows)
    metadata["excluded_duplicate_or_translation_resources"] = [resource_summary(item) for item in excluded]
    metadata["resource_selection_contract"] = "One English XLSX resource per reporting year and compliance category; alternate formats and translations are excluded."
    payload["metadata"] = metadata
    payload["toronto_rows"] = rows
    write_json(path, payload)
    return {"toronto_row_count": len(rows), "resource_count": len(selected), "excluded_resource_count": len(excluded)}


def parse_resource(resource: dict[str, Any]) -> list[dict[str, Any]]:
    url = str(resource.get("url") or "")
    if not url:
        raise RuntimeError(f"Ontario resource has no URL: {resource_summary(resource)}")
    data = request_bytes(url)
    fmt = str(resource.get("format") or "").upper()
    if fmt == "XLSX":
        return xlsx_rows(data)
    if fmt == "CSV":
        return csv_rows(data)
    raise RuntimeError(f"Unsupported Ontario format {fmt!r}")


def build_source(spec: dict[str, Any], output_dir: Path, retrieved_at: str) -> dict[str, Any]:
    package = package_show(spec["package"])
    licence = str(package.get("license_title") or "").strip()
    normalized_licence = normalize_licence_label(licence)
    if "open government licence" not in normalized_licence or "ontario" not in normalized_licence:
        raise RuntimeError(
            f"Ontario source {spec['key']} is not explicitly under Open Government Licence - Ontario: {licence!r}"
        )
    selected = select_resources(package, spec["mode"])
    if not selected:
        raise RuntimeError(f"Ontario source {spec['key']} had no selected tabular resources")

    all_toronto_rows: list[dict[str, Any]] = []
    resource_stats = []
    for resource in selected:
        rows = parse_resource(resource)
        toronto_rows = []
        bases: dict[str, int] = {}
        for row in rows:
            is_toronto, basis = row_is_toronto(row)
            bases[basis] = bases.get(basis, 0) + 1
            if is_toronto:
                enriched = dict(row)
                enriched["_towersignal_source_resource_id"] = resource.get("id")
                enriched["_towersignal_source_resource_name"] = resource.get("name")
                enriched["_towersignal_source_year"] = resource_year(resource)
                enriched["_towersignal_toronto_scope_basis"] = basis
                toronto_rows.append(enriched)
        all_toronto_rows.extend(toronto_rows)
        resource_stats.append({
            **resource_summary(resource),
            "source_row_count": len(rows),
            "toronto_row_count": len(toronto_rows),
            "scope_basis_counts": bases,
        })

    payload = {
        "metadata": {
            "key": spec["key"],
            "title": package.get("title"),
            "package_name": package.get("name"),
            "package_id": package.get("id"),
            "catalogue_url": f"https://data.ontario.ca/dataset/{package.get('name')}",
            "license": licence,
            "license_gate": "Catalogue licence label must normalize to contain both 'Open Government Licence' and 'Ontario'.",
            "retrieved_at": retrieved_at,
            "selection_mode": spec["mode"],
            "resource_count": len(selected),
            "toronto_row_count": len(all_toronto_rows),
            "resources": resource_stats,
            "scope_contract": "Exact Toronto/Etobicoke/North York/Scarborough/East York/York city-or-municipality field; fallback only to a Toronto M-postal code contained in an address field.",
            "absence": "No row is not evidence that a property/facility has no relevant equipment, activity, or compliance history.",
        },
        "toronto_rows": all_toronto_rows,
    }
    write_json(output_dir / "open_licensed" / f"{spec['key']}.json", payload)
    return payload["metadata"]


def write_json(path: Path, payload: Any, *, pretty: bool = False) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(payload, indent=2 if pretty else None, separators=None if pretty else (",", ":"), ensure_ascii=False, default=str),
        encoding="utf-8",
    )


def replace_inventory(inventory: dict[str, Any], source: dict[str, Any]) -> None:
    sources = inventory.setdefault("open_licensed_sources", [])
    sources[:] = [item for item in sources if item.get("key") != source.get("key")]
    sources.append(source)


def build(output_dir: Path) -> dict[str, Any]:
    inventory_path = output_dir / "source_inventory.json"
    inventory = json.loads(inventory_path.read_text(encoding="utf-8"))
    retrieved_at = utc_now()
    results = {}
    for spec in SOURCES:
        metadata = build_source(spec, output_dir, retrieved_at)
        replace_inventory(inventory, metadata)
        results[spec["key"]] = {
            "toronto_row_count": metadata["toronto_row_count"],
            "resource_count": metadata["resource_count"],
        }
    inventory["generated_at"] = retrieved_at
    inventory["ontario_open_building_environment_pull"] = results
    write_json(inventory_path, inventory, pretty=True)
    print(json.dumps(results, indent=2))
    return results


def main() -> None:
    parser = argparse.ArgumentParser(description="Pull Ontario open building/environment data for Toronto")
    parser.add_argument("--output", type=Path, default=ROOT / "data/toronto/warehouse/current")
    parser.add_argument("--normalize-existing", action="store_true", help="Normalize the persisted environmental report without network access")
    args = parser.parse_args()
    if args.normalize_existing:
        result = normalize_existing_environmental_report(args.output / "open_licensed" / "ontario_environmental_compliance_reports.json")
        print(json.dumps(result, indent=2))
        return
    build(args.output)


if __name__ == "__main__":
    main()
