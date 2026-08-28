from __future__ import annotations

import argparse
import csv
import io
import json
import re
import time
from pathlib import Path
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.parse import urlencode
from urllib.request import Request, urlopen

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
TORONTO_CKAN = "https://ckan0.cf.opendata.inter.prod-toronto.ca/api/3/action"
CHEMTRAC_PACKAGE = "chemical-tracking-chemtrac"
NOTICES_URL = "https://secure.toronto.ca/nm/notices.json"
USER_AGENT = "TowerSignal-Toronto-Warehouse/0.1 (+https://github.com/JeremyHennessy/TowerSignal)"
TORONTO_LICENSE = "Open Government Licence - Toronto"

STREET_SUFFIXES = {
    "STREET": "ST", "ST": "ST", "AVENUE": "AVE", "AVE": "AVE", "ROAD": "RD", "RD": "RD",
    "BOULEVARD": "BLVD", "BLVD": "BLVD", "DRIVE": "DR", "DR": "DR", "COURT": "CT", "CT": "CT",
    "CRESCENT": "CRES", "CRES": "CRES", "PARKWAY": "PKWY", "PKWY": "PKWY", "TRAIL": "TRL", "TRL": "TRL",
    "PLACE": "PL", "PL": "PL", "TERRACE": "TER", "TER": "TER", "HIGHWAY": "HWY", "HWY": "HWY",
}
DIRECTIONS = {"EAST": "E", "WEST": "W", "NORTH": "N", "SOUTH": "S", "E": "E", "W": "W", "N": "N", "S": "S"}


def request_bytes(url: str, *, timeout: int = 180, retries: int = 4) -> bytes:
    last_error: Exception | None = None
    for attempt in range(retries):
        try:
            request = Request(url, headers={"User-Agent": USER_AGENT, "Accept": "application/json,*/*;q=0.8"})
            with urlopen(request, timeout=timeout) as response:
                return response.read()
        except (HTTPError, URLError, TimeoutError) as exc:
            last_error = exc
            if attempt + 1 < retries:
                time.sleep(2 * (attempt + 1))
    raise RuntimeError(f"Failed to retrieve {url}: {last_error}")


def request_json(url: str) -> Any:
    return json.loads(request_bytes(url).decode("utf-8"))


def ckan(action: str, params: dict[str, Any]) -> dict[str, Any]:
    payload = request_json(f"{TORONTO_CKAN}/{action}?{urlencode(params)}")
    if not isinstance(payload, dict) or payload.get("success") is not True or not isinstance(payload.get("result"), dict):
        raise RuntimeError(f"Toronto CKAN {action} failed: {payload!r}")
    return payload["result"]


def canonical_address(value: Any) -> str | None:
    if value in (None, ""):
        return None
    text = str(value).upper()
    text = re.sub(r"\b[A-Z]\d[A-Z]\s*\d[A-Z]\d\b.*$", "", text)
    text = text.split(",", 1)[0]
    text = re.sub(r"[^A-Z0-9]+", " ", text)
    tokens = [STREET_SUFFIXES.get(token, DIRECTIONS.get(token, token)) for token in text.split()]
    result = " ".join(tokens).strip()
    return result or None


def read_json(path: Path) -> dict[str, Any]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(payload, dict):
        raise RuntimeError(f"Expected JSON object: {path}")
    return payload


def write_json(path: Path, payload: Any, *, pretty: bool = False) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(payload, indent=2 if pretty else None, separators=None if pretty else (",", ":"), ensure_ascii=False, default=str),
        encoding="utf-8",
    )


def ckan_datastore_rows(resource_id: str) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    offset = 0
    limit = 5000
    total: int | None = None
    while total is None or offset < total:
        result = ckan("datastore_search", {"resource_id": resource_id, "limit": limit, "offset": offset})
        total = int(result.get("total") or 0) if total is None else total
        batch = result.get("records") or []
        if not isinstance(batch, list):
            raise RuntimeError("ChemTRAC datastore records were not a list")
        rows.extend(item for item in batch if isinstance(item, dict))
        if not batch:
            break
        offset += len(batch)
    return rows


def tabular_rows(resource: dict[str, Any]) -> list[dict[str, Any]]:
    if resource.get("datastore_active"):
        return ckan_datastore_rows(str(resource["id"]))
    url = str(resource.get("url") or "")
    fmt = str(resource.get("format") or "").upper()
    data = request_bytes(url)
    if fmt == "CSV" or url.lower().endswith(".csv"):
        return [dict(row) for row in csv.DictReader(io.StringIO(data.decode("utf-8-sig", errors="replace")))]
    if fmt in {"XLSX", "XLS"} or url.lower().endswith(".xlsx"):
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        output: list[dict[str, Any]] = []
        for sheet in workbook.worksheets:
            iterator = sheet.iter_rows(values_only=True)
            header = None
            for row in iterator:
                values = list(row)
                if sum(value is not None and str(value).strip() for value in values) >= 3:
                    header = [re.sub(r"\s+", " ", str(value or f"column_{idx+1}")).strip() for idx, value in enumerate(values)]
                    break
            if not header:
                continue
            for row in iterator:
                values = list(row)
                if not any(value is not None and str(value).strip() for value in values):
                    continue
                output.append({header[idx]: values[idx] if idx < len(values) else None for idx in range(len(header))})
        return output
    raise RuntimeError(f"Unsupported ChemTRAC historical resource: {resource.get('name')} / {fmt}")


def year_from_resource(resource: dict[str, Any]) -> int | None:
    text = " ".join(str(resource.get(key) or "") for key in ("name", "description", "url"))
    years = re.findall(r"\b(20\d{2})\b", text)
    return max((int(year) for year in years), default=None)


def pull_chemtrac_history(output_dir: Path) -> dict[str, Any]:
    package = ckan("package_show", {"id": CHEMTRAC_PACKAGE})
    resources = [item for item in (package.get("resources") or []) if isinstance(item, dict)]
    historical = []
    stats = []
    for resource in resources:
        year = year_from_resource(resource)
        name = str(resource.get("name") or "")
        if year is None or "chemtrac" not in name.lower():
            continue
        fmt = str(resource.get("format") or "").upper()
        if not resource.get("datastore_active") and fmt not in {"CSV", "XLSX"}:
            continue
        rows = tabular_rows(resource)
        for row in rows:
            enriched = dict(row)
            enriched["_towersignal_reporting_year"] = year
            enriched["_towersignal_source_resource_id"] = resource.get("id")
            historical.append(enriched)
        stats.append({
            "year": year,
            "resource_id": resource.get("id"),
            "name": name,
            "format": fmt,
            "last_modified": resource.get("last_modified"),
            "row_count": len(rows),
        })
    if not stats:
        raise RuntimeError("No historical ChemTRAC annual resources were parsed")
    stats.sort(key=lambda item: (item["year"], item["name"]))
    payload = {
        "metadata": {
            "key": "chemtrac_history",
            "title": package.get("title"),
            "portal_url": f"https://open.toronto.ca/dataset/{package.get('name')}/",
            "license": TORONTO_LICENSE,
            "resource_count": len(stats),
            "row_count": len(historical),
            "resources": stats,
            "known_gap": "Toronto Open Data notes ChemTRAC data from 2019 through 2023 is unavailable due to COVID-19 related reporting interruptions.",
        },
        "rows": historical,
    }
    write_json(output_dir / "open_licensed" / "chemtrac_history.json", payload)
    return payload["metadata"]


def list_of_dicts(payload: Any) -> list[dict[str, Any]]:
    if isinstance(payload, list):
        return [item for item in payload if isinstance(item, dict)]
    if isinstance(payload, dict):
        for key in ("notices", "results", "data", "items"):
            value = payload.get(key)
            if isinstance(value, list):
                return [item for item in value if isinstance(item, dict)]
        list_values = [value for value in payload.values() if isinstance(value, list) and value and isinstance(value[0], dict)]
        if len(list_values) == 1:
            return [item for item in list_values[0] if isinstance(item, dict)]
    raise RuntimeError("Could not identify notice list in Toronto Public Notices JSON")


def flatten_strings(value: Any) -> list[str]:
    if isinstance(value, str):
        return [value]
    if isinstance(value, dict):
        output = []
        for item in value.values():
            output.extend(flatten_strings(item))
        return output
    if isinstance(value, list):
        output = []
        for item in value:
            output.extend(flatten_strings(item))
        return output
    return []


def notice_is_planning(notice: dict[str, Any]) -> bool:
    if notice.get("planningApplicationNumbers"):
        return True
    text = " | ".join(flatten_strings(notice.get("topics")) + [str(notice.get("title") or "")]).lower()
    return "planning" in text or "zoning" in text or "site plan" in text or "development application" in text


def notice_addresses(notice: dict[str, Any]) -> set[str]:
    addresses: set[str] = set()
    for text in flatten_strings(notice.get("addressList")):
        candidate = canonical_address(text)
        if candidate and re.search(r"\d", candidate):
            addresses.add(candidate)
    return addresses


def pull_public_notices(poc_dir: Path, output_dir: Path) -> dict[str, Any]:
    raw = request_json(NOTICES_URL)
    notices = list_of_dicts(raw)
    planning = [notice for notice in notices if notice_is_planning(notice)]
    poc_properties = read_json(poc_dir / "properties.json").get("properties") or []
    poc_addresses: dict[str, list[dict[str, Any]]] = {}
    for item in poc_properties:
        if not isinstance(item, dict):
            continue
        address = canonical_address(item.get("address"))
        if address:
            poc_addresses.setdefault(address, []).append(item)

    matches = []
    matched_addresses: set[str] = set()
    matched_confirmed: set[str] = set()
    for notice in planning:
        overlap = sorted(notice_addresses(notice) & set(poc_addresses))
        if not overlap:
            continue
        for address in overlap:
            matched_addresses.add(address)
            if any(item.get("tower_status") == "CONFIRMED" for item in poc_addresses[address]):
                matched_confirmed.add(address)
        matches.append({
            "noticeId": notice.get("noticeId"),
            "title": notice.get("title"),
            "noticeDate": notice.get("noticeDate"),
            "planningApplicationNumbers": notice.get("planningApplicationNumbers"),
            "topics": notice.get("topics"),
            "addressList": notice.get("addressList"),
            "backgroundInformationList": notice.get("backgroundInformationList"),
            "otherReferenceList": notice.get("otherReferenceList"),
            "contact": notice.get("contact"),
            "matched_canonical_addresses": overlap,
            "matched_property_keys": sorted({
                str(item.get("property_key"))
                for address in overlap
                for item in poc_addresses[address]
                if item.get("property_key")
            }),
        })

    payload = {
        "metadata": {
            "key": "toronto_public_notices",
            "source_url": NOTICES_URL,
            "documentation_url": "https://secure.toronto.ca/nm/opendata.do",
            "license": TORONTO_LICENSE,
            "licence_basis": "City Clerk publishes this as an Open Data API; use follows the City of Toronto Open Government Licence for City Open Data information.",
            "source_notice_count": len(notices),
            "planning_notice_count": len(planning),
            "matched_planning_notice_count": len(matches),
            "matched_poc_address_count": len(matched_addresses),
            "matched_confirmed_tower_address_count": len(matched_confirmed),
            "match_basis": "EXACT_CANONICAL_ADDRESS_FROM_NOTICE_ADDRESS_LIST_TO_POC_ADDRESS",
            "background_documents": "BackgroundInformationList is retained as source document context; document contents are not automatically treated as cooling-tower evidence.",
        },
        "planning_notices": planning,
        "poc_matches": matches,
    }
    write_json(output_dir / "open_licensed" / "toronto_public_notices.json", payload)
    return payload["metadata"]


def replace_inventory(inventory: dict[str, Any], source: dict[str, Any]) -> None:
    sources = inventory.setdefault("open_licensed_sources", [])
    sources[:] = [item for item in sources if item.get("key") != source.get("key")]
    sources.append(source)


def build(poc_dir: Path, output_dir: Path) -> dict[str, Any]:
    inventory_path = output_dir / "source_inventory.json"
    inventory = read_json(inventory_path)
    chemtrac = pull_chemtrac_history(output_dir)
    notices = pull_public_notices(poc_dir, output_dir)
    replace_inventory(inventory, chemtrac)
    replace_inventory(inventory, notices)
    result = {
        "chemtrac_history": {"row_count": chemtrac["row_count"], "resource_count": chemtrac["resource_count"]},
        "public_notices": {
            "source_notice_count": notices["source_notice_count"],
            "planning_notice_count": notices["planning_notice_count"],
            "matched_poc_address_count": notices["matched_poc_address_count"],
            "matched_confirmed_tower_address_count": notices["matched_confirmed_tower_address_count"],
        },
    }
    inventory["historical_and_notice_pull"] = result
    write_json(inventory_path, inventory, pretty=True)
    print(json.dumps(result, indent=2))
    return result


def main() -> None:
    parser = argparse.ArgumentParser(description="Pull Toronto historical ChemTRAC and Public Notices open data")
    parser.add_argument("--poc", type=Path, default=ROOT / "data/toronto/poc/current")
    parser.add_argument("--warehouse", type=Path, default=ROOT / "data/toronto/warehouse/current")
    args = parser.parse_args()
    build(args.poc, args.warehouse)


if __name__ == "__main__":
    main()
